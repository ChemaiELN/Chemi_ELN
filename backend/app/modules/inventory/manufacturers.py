"""Inventory – Manufacturers endpoints."""
from typing import Any, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from pydantic import ValidationError
from sqlalchemy.orm import Session

from app.dependencies import get_current_user, get_db
from app.models.inventory import InvManufacturer
from app.schemas.inventory import (
    ManufacturerCreate,
    ManufacturerListOut,
    ManufacturerOut,
    ManufacturerUpdate,
    ManufacturerUploadResult,
)

router = APIRouter(prefix="/inventory/manufacturers", tags=["inventory-manufacturers"])

# Whitelist of columns the Manufacturers table UI is allowed to sort by.
SORTABLE_COLUMNS = {
    "code": InvManufacturer.code,
    "name": InvManufacturer.name,
    "country": InvManufacturer.country,
    "contact_person": InvManufacturer.contact_person,
    "email": InvManufacturer.email,
    "phone": InvManufacturer.phone,
    "is_active": InvManufacturer.is_active,
}


@router.get("", response_model=ManufacturerListOut)
def list_manufacturers(
    search: Optional[str] = Query(None),
    active_only: bool = Query(False),
    skip: int = Query(0, ge=0),
    limit: int = Query(50, le=200),
    sort_by: Optional[str] = Query(None),
    sort_dir: str = Query("asc"),
    db: Session = Depends(get_db),
    _: Any = Depends(get_current_user),
):
    q = db.query(InvManufacturer)
    if active_only:
        q = q.filter(InvManufacturer.is_active.is_(True))
    if search:
        term = f"%{search}%"
        q = q.filter(
            InvManufacturer.name.ilike(term)
            | InvManufacturer.code.ilike(term)
            | InvManufacturer.country.ilike(term)
            | InvManufacturer.contact_person.ilike(term)
            | InvManufacturer.email.ilike(term)
            | InvManufacturer.phone.ilike(term)
        )
    total = q.count()
    sort_col = SORTABLE_COLUMNS.get(sort_by, InvManufacturer.name)
    order_clause = sort_col.desc() if sort_dir == "desc" else sort_col.asc()
    items = q.order_by(order_clause).offset(skip).limit(limit).all()
    return {"items": items, "total": total}


@router.post("", response_model=ManufacturerOut, status_code=201)
def create_manufacturer(
    body: ManufacturerCreate,
    db: Session = Depends(get_db),
    _: Any = Depends(get_current_user),
):
    if db.query(InvManufacturer).filter_by(code=body.code).first():
        raise HTTPException(409, f"Manufacturer code '{body.code}' already exists.")
    row = InvManufacturer(**body.model_dump())
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.get("/{manufacturer_id}", response_model=ManufacturerOut)
def get_manufacturer(
    manufacturer_id: int,
    db: Session = Depends(get_db),
    _: Any = Depends(get_current_user),
):
    row = db.get(InvManufacturer, manufacturer_id)
    if not row:
        raise HTTPException(404, "Manufacturer not found.")
    return row


@router.patch("/{manufacturer_id}", response_model=ManufacturerOut)
def update_manufacturer(
    manufacturer_id: int,
    body: ManufacturerUpdate,
    db: Session = Depends(get_db),
    _: Any = Depends(get_current_user),
):
    row = db.get(InvManufacturer, manufacturer_id)
    if not row:
        raise HTTPException(404, "Manufacturer not found.")
    for k, v in body.model_dump(exclude_none=True).items():
        setattr(row, k, v)
    db.commit()
    db.refresh(row)
    return row


@router.delete("/{manufacturer_id}/deactivate", response_model=ManufacturerOut)
def deactivate_manufacturer(
    manufacturer_id: int,
    db: Session = Depends(get_db),
    _: Any = Depends(get_current_user),
):
    row = db.get(InvManufacturer, manufacturer_id)
    if not row:
        raise HTTPException(404, "Manufacturer not found.")
    row.is_active = False
    db.commit()
    db.refresh(row)
    return row


@router.patch("/{manufacturer_id}/toggle", response_model=ManufacturerOut)
def toggle_manufacturer(
    manufacturer_id: int,
    db: Session = Depends(get_db),
    _: Any = Depends(get_current_user),
):
    row = db.get(InvManufacturer, manufacturer_id)
    if not row:
        raise HTTPException(404, "Manufacturer not found.")
    row.is_active = not row.is_active
    db.commit()
    db.refresh(row)
    out = ManufacturerOut.model_validate(row)
    out.message = f"{row.name} {'activated' if row.is_active else 'deactivated'}."
    return out


# ── Bulk Excel upload ──────────────────────────────────────────────────────────
# Column order must match master_templates.py's TEMPLATE_SHEETS["manufacturers"]["headers"].
UPLOAD_COLUMNS = [
    "code", "name", "country", "contact_person",
    "email", "phone", "website", "address",
]


@router.post("/upload", response_model=ManufacturerUploadResult)
def upload_manufacturers(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: Any = Depends(get_current_user),
):
    from openpyxl import load_workbook

    try:
        wb = load_workbook(file.file, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(400, f"Could not read Excel file: {exc}")
    ws = wb.active

    created, skipped, errors = 0, 0, []
    seen_codes: dict[str, int] = {}  # code -> first row number that used it, within this file

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(c is None for c in row):
            continue
        values = dict(zip(UPLOAD_COLUMNS, (row + (None,) * len(UPLOAD_COLUMNS))[:len(UPLOAD_COLUMNS)]))

        try:
            body = ManufacturerCreate(**{k: v for k, v in values.items() if v is not None})
        except ValidationError as exc:
            msgs = "; ".join(f"{e['loc'][0]}: {e['msg']}" for e in exc.errors())
            errors.append(f"Row {i}: {msgs}")
            skipped += 1
            continue

        if body.code in seen_codes:
            errors.append(f"Row {i}: Code '{body.code}' duplicates row {seen_codes[body.code]} in this file.")
            skipped += 1
            continue
        if db.query(InvManufacturer).filter(InvManufacturer.code == body.code).first():
            errors.append(f"Row {i}: Code '{body.code}' is already used by another manufacturer.")
            skipped += 1
            continue

        db.add(InvManufacturer(**body.model_dump()))
        seen_codes[body.code] = i
        created += 1

    db.commit()
    return ManufacturerUploadResult(created=created, skipped=skipped, errors=errors)
