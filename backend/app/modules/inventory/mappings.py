"""Inventory – Material-Manufacturer Mappings + DSD file upload/download."""
import os
from typing import Any, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse
from pydantic import ValidationError
from sqlalchemy.orm import Session

from app.dependencies import get_current_user, get_db
from app.models.inventory import InvManufacturerMapping, InvMaterial, InvManufacturer
from app.schemas.inventory import MappingCreate, MappingListOut, MappingOut, MappingUpdate, MappingUploadResult
from app.shared.files import ALLOWED_DOC_EXTS, delete_file, save_upload, validate_upload

router = APIRouter(prefix="/inventory/mappings", tags=["inventory-mappings"])

# Whitelist of columns the Mappings table UI is allowed to sort by.
SORTABLE_COLUMNS = {
    "catalogue_no": InvManufacturerMapping.catalogue_no,
    "technical_grade": InvManufacturerMapping.technical_grade,
    "lead_time_days": InvManufacturerMapping.lead_time_days,
    "min_order_qty": InvManufacturerMapping.min_order_qty,
    "created_at": InvManufacturerMapping.created_at,
}


@router.get("", response_model=MappingListOut)
def list_mappings(
    search: Optional[str] = Query(None),
    material_id: Optional[int] = Query(None),
    manufacturer_id: Optional[int] = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(50, le=200),
    sort_by: Optional[str] = Query(None),
    sort_dir: str = Query("desc"),
    db: Session = Depends(get_db),
    _: Any = Depends(get_current_user),
):
    q = db.query(InvManufacturerMapping)
    if material_id is not None:
        q = q.filter(InvManufacturerMapping.material_id == material_id)
    if manufacturer_id is not None:
        q = q.filter(InvManufacturerMapping.manufacturer_id == manufacturer_id)
    if search:
        term = f"%{search}%"
        q = (
            q.outerjoin(InvMaterial, InvManufacturerMapping.material_id == InvMaterial.id)
            .outerjoin(InvManufacturer, InvManufacturerMapping.manufacturer_id == InvManufacturer.id)
            .filter(
                InvManufacturerMapping.catalogue_no.ilike(term)
                | InvManufacturerMapping.technical_grade.ilike(term)
                | InvMaterial.name.ilike(term)
                | InvManufacturer.name.ilike(term)
            )
        )
    total = q.count()
    sort_col = SORTABLE_COLUMNS.get(sort_by, InvManufacturerMapping.id)
    order_clause = sort_col.desc() if sort_dir == "desc" else sort_col.asc()
    items = q.order_by(order_clause).offset(skip).limit(limit).all()
    return {"items": items, "total": total}


@router.post("", response_model=MappingOut, status_code=201)
def create_mapping(
    body: MappingCreate,
    db: Session = Depends(get_db),
    _: Any = Depends(get_current_user),
):
    exists = (
        db.query(InvManufacturerMapping)
        .filter_by(material_id=body.material_id, manufacturer_id=body.manufacturer_id)
        .first()
    )
    if exists:
        raise HTTPException(409, "Mapping for this material+manufacturer already exists.")
    row = InvManufacturerMapping(**body.model_dump())
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.get("/{mapping_id}", response_model=MappingOut)
def get_mapping(
    mapping_id: int,
    db: Session = Depends(get_db),
    _: Any = Depends(get_current_user),
):
    row = db.get(InvManufacturerMapping, mapping_id)
    if not row:
        raise HTTPException(404, "Mapping not found.")
    return row


@router.patch("/{mapping_id}", response_model=MappingOut)
def update_mapping(
    mapping_id: int,
    body: MappingUpdate,
    db: Session = Depends(get_db),
    _: Any = Depends(get_current_user),
):
    row = db.get(InvManufacturerMapping, mapping_id)
    if not row:
        raise HTTPException(404, "Mapping not found.")
    for k, v in body.model_dump(exclude_none=True).items():
        setattr(row, k, v)
    db.commit()
    db.refresh(row)
    return row


@router.delete("/{mapping_id}", status_code=204)
def delete_mapping(
    mapping_id: int,
    db: Session = Depends(get_db),
    _: Any = Depends(get_current_user),
):
    row = db.get(InvManufacturerMapping, mapping_id)
    if not row:
        raise HTTPException(404, "Mapping not found.")
    delete_file(row.dsd_file_path)
    db.delete(row)
    db.commit()


# ── DSD file upload / download ─────────────────────────────────────────────────
@router.post("/{mapping_id}/dsd", response_model=MappingOut)
async def upload_dsd(
    mapping_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: Any = Depends(get_current_user),
):
    row = db.get(InvManufacturerMapping, mapping_id)
    if not row:
        raise HTTPException(404, "Mapping not found.")
    validate_upload(file, allowed_exts=ALLOWED_DOC_EXTS)
    if row.dsd_file_path:
        delete_file(row.dsd_file_path)
    row.dsd_file_path = await save_upload(file, subdir="dsd")
    db.commit()
    db.refresh(row)
    return row


@router.get("/{mapping_id}/dsd/download")
def download_dsd(
    mapping_id: int,
    db: Session = Depends(get_db),
    _: Any = Depends(get_current_user),
):
    row = db.get(InvManufacturerMapping, mapping_id)
    if not row or not row.dsd_file_path:
        raise HTTPException(404, "No DSD file attached to this mapping.")
    if not os.path.exists(row.dsd_file_path):
        raise HTTPException(404, "DSD file missing from disk.")
    return FileResponse(
        row.dsd_file_path,
        media_type="application/octet-stream",
        filename=os.path.basename(row.dsd_file_path),
    )


@router.delete("/{mapping_id}/dsd", response_model=MappingOut)
def delete_dsd(
    mapping_id: int,
    db: Session = Depends(get_db),
    _: Any = Depends(get_current_user),
):
    row = db.get(InvManufacturerMapping, mapping_id)
    if not row:
        raise HTTPException(404, "Mapping not found.")
    delete_file(row.dsd_file_path)
    row.dsd_file_path = None
    db.commit()
    db.refresh(row)
    return row


# ── Bulk Excel upload ──────────────────────────────────────────────────────────
# Column order must match master_templates.py's TEMPLATE_SHEETS["mappings"]["headers"].
UPLOAD_COLUMNS = [
    "material_code", "manufacturer_code", "catalogue_no",
    "technical_grade", "lead_time_days", "min_order_qty",
]


@router.post("/upload", response_model=MappingUploadResult)
def upload_mappings(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: Any = Depends(get_current_user),
):
    from openpyxl import load_workbook

    try:
        wb = load_workbook(file.file, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(400, f"Could not read Excel file: {exc}")
    ws = wb.active

    created, skipped, errors = 0, 0, []
    seen_pairs: dict[tuple[int, int], int] = {}  # (material_id, manufacturer_id) -> first row number

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(c is None for c in row):
            continue
        values = dict(zip(UPLOAD_COLUMNS, (row + (None,) * len(UPLOAD_COLUMNS))[:len(UPLOAD_COLUMNS)]))
        material_code = values.pop("material_code", None)
        manufacturer_code = values.pop("manufacturer_code", None)

        if not material_code or not str(material_code).strip():
            errors.append(f"Row {i}: Material Code is required.")
            skipped += 1
            continue
        if not manufacturer_code or not str(manufacturer_code).strip():
            errors.append(f"Row {i}: Manufacturer Code is required.")
            skipped += 1
            continue

        material = db.query(InvMaterial).filter(InvMaterial.code == str(material_code).strip()).first()
        if not material:
            errors.append(f"Row {i}: Material Code '{material_code}' not found.")
            skipped += 1
            continue
        manufacturer = db.query(InvManufacturer).filter(InvManufacturer.code == str(manufacturer_code).strip()).first()
        if not manufacturer:
            errors.append(f"Row {i}: Manufacturer Code '{manufacturer_code}' not found.")
            skipped += 1
            continue

        try:
            body = MappingCreate(
                material_id=material.id,
                manufacturer_id=manufacturer.id,
                **{k: v for k, v in values.items() if v is not None},
            )
        except ValidationError as exc:
            msgs = "; ".join(f"{e['loc'][0]}: {e['msg']}" for e in exc.errors())
            errors.append(f"Row {i}: {msgs}")
            skipped += 1
            continue

        pair = (body.material_id, body.manufacturer_id)
        if pair in seen_pairs:
            errors.append(f"Row {i}: duplicates row {seen_pairs[pair]} in this file (same material + manufacturer).")
            skipped += 1
            continue
        if db.query(InvManufacturerMapping).filter_by(material_id=body.material_id, manufacturer_id=body.manufacturer_id).first():
            errors.append(f"Row {i}: mapping for '{material_code}' + '{manufacturer_code}' already exists.")
            skipped += 1
            continue

        db.add(InvManufacturerMapping(**body.model_dump()))
        seen_pairs[pair] = i
        created += 1

    db.commit()
    return MappingUploadResult(created=created, skipped=skipped, errors=errors)
