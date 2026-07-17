"""Inventory – Materials, ChemicalProps, FormulationProps endpoints."""
import datetime
from typing import Any, Optional
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from pydantic import ValidationError
from sqlalchemy.orm import Session

from app.dependencies import get_current_user, get_db
from app.models.inventory import (
    InvConsumableType,
    InvMaterialChemicalProps,
    InvMaterialFormulationProps,
    InvMaterial,
    InvMaterialCodeCounter,
)
from app.schemas.inventory import (
    ChemicalPropsOut,
    ChemicalPropsUpsert,
    FormulationPropsOut,
    FormulationPropsUpsert,
    MaterialCreate,
    MaterialListOut,
    MaterialOut,
    MaterialUpdate,
    MaterialUploadResult,
)

router = APIRouter(prefix="/inventory/materials", tags=["inventory-materials"])

CODE_PREFIX = "MAT"

# Whitelist of columns the Materials table UI is allowed to sort by.
SORTABLE_COLUMNS = {
    "code": InvMaterial.code,
    "name": InvMaterial.name,
    "material_type": InvMaterial.material_type,
    "cas_no": InvMaterial.cas_no,
    "molecular_formula": InvMaterial.molecular_formula,
    "mol_weight": InvMaterial.mol_weight,
    "storage_condition": InvMaterial.storage_condition,
    "hazard_class": InvMaterial.hazard_class,
    "is_active": InvMaterial.is_active,
}


def _seed_max_seq(db: Session, year: str) -> int:
    """Seed from the highest existing sequence in any material code for the year."""
    pattern = f"{CODE_PREFIX}/{year}/%"
    rows = db.query(InvMaterial.code).filter(InvMaterial.code.like(pattern)).all()
    max_seq = 10000
    for (code,) in rows:
        if code:
            try:
                s = int(code.split('/')[-1])
                if s > max_seq:
                    max_seq = s
            except (ValueError, IndexError):
                pass
    return max_seq


def _claim_next_seq(db: Session, year: str) -> int:
    """Atomically increment and return the next sequence for the year.
    SELECT FOR UPDATE prevents two concurrent requests getting the same number."""
    counter = (
        db.query(InvMaterialCodeCounter)
        .filter_by(year=year)
        .with_for_update()
        .first()
    )
    if counter is None:
        counter = InvMaterialCodeCounter(year=year, last_seq=_seed_max_seq(db, year))
        db.add(counter)
        db.flush()
    counter.last_seq += 1
    return counter.last_seq


@router.get("/next-code")
def next_code(
    db: Session = Depends(get_db),
    _: Any = Depends(get_current_user),
):
    """Preview-only: returns the likely next material code without claiming it."""
    year = datetime.datetime.utcnow().strftime('%y')
    counter = db.query(InvMaterialCodeCounter).filter_by(year=year).first()
    next_seq = (counter.last_seq if counter else _seed_max_seq(db, year)) + 1
    return {"code": f"{CODE_PREFIX}/{year}/{next_seq}"}


@router.get("", response_model=MaterialListOut)
def list_materials(
    search: Optional[str] = Query(None),
    material_type: Optional[str] = Query(None),
    consumable_type_id: Optional[int] = Query(None),
    department_id: Optional[UUID] = Query(None),
    active_only: bool = Query(False),
    skip: int = Query(0, ge=0),
    limit: int = Query(50, le=200),
    sort_by: Optional[str] = Query(None),
    sort_dir: str = Query("asc"),
    db: Session = Depends(get_db),
    _: Any = Depends(get_current_user),
):
    q = db.query(InvMaterial)
    if active_only:
        q = q.filter(InvMaterial.is_active.is_(True))
    if material_type is not None:
        q = q.filter(InvMaterial.material_type.ilike(material_type))
    if consumable_type_id is not None:
        q = q.filter(InvMaterial.consumable_type_id == consumable_type_id)
    if department_id is not None:
        q = q.filter(InvMaterial.department_id == department_id)
    if search:
        term = f"%{search}%"
        q = q.filter(
            InvMaterial.name.ilike(term)
            | InvMaterial.code.ilike(term)
            | InvMaterial.cas_no.ilike(term)
            | InvMaterial.molecular_formula.ilike(term)
            | InvMaterial.hazard_class.ilike(term)
            | InvMaterial.material_type.ilike(term)
        )
    total = q.count()
    sort_col = SORTABLE_COLUMNS.get(sort_by, InvMaterial.code)
    order_clause = sort_col.desc() if sort_dir == "desc" else sort_col.asc()
    items = q.order_by(order_clause).offset(skip).limit(limit).all()
    return {"items": items, "total": total}


@router.post("", response_model=MaterialOut, status_code=201)
def create_material(
    body: MaterialCreate,
    db: Session = Depends(get_db),
    _: Any = Depends(get_current_user),
):
    if db.query(InvMaterial).filter(InvMaterial.cas_no == body.cas_no).first():
        raise HTTPException(409, f"CAS No '{body.cas_no}' is already used by another material.")
    data = body.model_dump(exclude={"code"})
    year = datetime.datetime.utcnow().strftime('%y')
    seq = _claim_next_seq(db, year)
    row = InvMaterial(code=f"{CODE_PREFIX}/{year}/{seq}", **data)
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@router.get("/{material_id}", response_model=MaterialOut)
def get_material(
    material_id: int,
    db: Session = Depends(get_db),
    _: Any = Depends(get_current_user),
):
    row = db.get(InvMaterial, material_id)
    if not row:
        raise HTTPException(404, "Material not found.")
    return row


@router.patch("/{material_id}", response_model=MaterialOut)
def update_material(
    material_id: int,
    body: MaterialUpdate,
    db: Session = Depends(get_db),
    _: Any = Depends(get_current_user),
):
    row = db.get(InvMaterial, material_id)
    if not row:
        raise HTTPException(404, "Material not found.")
    if body.cas_no is not None and body.cas_no != row.cas_no:
        if db.query(InvMaterial).filter(InvMaterial.cas_no == body.cas_no, InvMaterial.id != material_id).first():
            raise HTTPException(409, f"CAS No '{body.cas_no}' is already used by another material.")
    for k, v in body.model_dump(exclude_none=True).items():
        setattr(row, k, v)
    db.commit()
    db.refresh(row)
    return row


@router.delete("/{material_id}/deactivate", response_model=MaterialOut)
def deactivate_material(
    material_id: int,
    db: Session = Depends(get_db),
    _: Any = Depends(get_current_user),
):
    row = db.get(InvMaterial, material_id)
    if not row:
        raise HTTPException(404, "Material not found.")
    row.is_active = False
    db.commit()
    db.refresh(row)
    return row


@router.patch("/{material_id}/toggle", response_model=MaterialOut)
def toggle_material(
    material_id: int,
    db: Session = Depends(get_db),
    _: Any = Depends(get_current_user),
):
    row = db.get(InvMaterial, material_id)
    if not row:
        raise HTTPException(404, "Material not found.")
    row.is_active = not row.is_active
    db.commit()
    db.refresh(row)
    out = MaterialOut.model_validate(row)
    out.message = f"{row.name} {'activated' if row.is_active else 'deactivated'}."
    return out


# ── Chemical Props ─────────────────────────────────────────────────────────────
@router.get("/{material_id}/chemical-props", response_model=Optional[ChemicalPropsOut])
def get_chemical_props(
    material_id: int,
    db: Session = Depends(get_db),
    _: Any = Depends(get_current_user),
):
    mat = db.get(InvMaterial, material_id)
    if not mat:
        raise HTTPException(404, "Material not found.")
    return mat.chemical_props


@router.put("/{material_id}/chemical-props", response_model=ChemicalPropsOut)
def upsert_chemical_props(
    material_id: int,
    body: ChemicalPropsUpsert,
    db: Session = Depends(get_db),
    _: Any = Depends(get_current_user),
):
    mat = db.get(InvMaterial, material_id)
    if not mat:
        raise HTTPException(404, "Material not found.")
    props = mat.chemical_props
    if props is None:
        props = InvMaterialChemicalProps(material_id=material_id)
        db.add(props)
    for k, v in body.model_dump(exclude_none=True).items():
        setattr(props, k, v)
    db.commit()
    db.refresh(props)
    return props


# ── Formulation Props ──────────────────────────────────────────────────────────
@router.get("/{material_id}/formulation-props", response_model=Optional[FormulationPropsOut])
def get_formulation_props(
    material_id: int,
    db: Session = Depends(get_db),
    _: Any = Depends(get_current_user),
):
    mat = db.get(InvMaterial, material_id)
    if not mat:
        raise HTTPException(404, "Material not found.")
    return mat.formulation_props


@router.put("/{material_id}/formulation-props", response_model=FormulationPropsOut)
def upsert_formulation_props(
    material_id: int,
    body: FormulationPropsUpsert,
    db: Session = Depends(get_db),
    _: Any = Depends(get_current_user),
):
    mat = db.get(InvMaterial, material_id)
    if not mat:
        raise HTTPException(404, "Material not found.")
    props = mat.formulation_props
    if props is None:
        props = InvMaterialFormulationProps(material_id=material_id)
        db.add(props)
    for k, v in body.model_dump(exclude_none=True).items():
        setattr(props, k, v)
    db.commit()
    db.refresh(props)
    return props


# ── Bulk Excel upload ──────────────────────────────────────────────────────────
# Column order must match master_templates.py's TEMPLATE_SHEETS["materials"]["headers"].
UPLOAD_COLUMNS = [
    "name", "material_type", "cas_no", "molecular_formula",
    "mol_weight", "storage_condition", "hazard_class",
    "consumable_type_name", "description",
]


@router.post("/upload", response_model=MaterialUploadResult)
def upload_materials(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: Any = Depends(get_current_user),
):
    from openpyxl import load_workbook

    try:
        wb = load_workbook(file.file, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(400, f"Could not read Excel file: {exc}")
    ws = wb.active

    created, skipped, errors = 0, 0, []
    year = datetime.datetime.utcnow().strftime('%y')
    seen_cas: dict[str, int] = {}  # cas_no -> first row number that used it, within this file

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(c is None for c in row):
            continue
        values = dict(zip(UPLOAD_COLUMNS, (row + (None,) * len(UPLOAD_COLUMNS))[:len(UPLOAD_COLUMNS)]))
        consumable_type_name = values.pop("consumable_type_name", None)

        mol_weight = values.get("mol_weight")
        if mol_weight is not None:
            values["mol_weight"] = str(mol_weight).strip() or None

        try:
            body = MaterialCreate(**{k: v for k, v in values.items() if v is not None})
        except ValidationError as exc:
            msgs = "; ".join(f"{e['loc'][0]}: {e['msg']}" for e in exc.errors())
            errors.append(f"Row {i}: {msgs}")
            skipped += 1
            continue

        if body.cas_no in seen_cas:
            errors.append(f"Row {i}: CAS No '{body.cas_no}' duplicates row {seen_cas[body.cas_no]} in this file.")
            skipped += 1
            continue
        if db.query(InvMaterial).filter(InvMaterial.cas_no == body.cas_no).first():
            errors.append(f"Row {i}: CAS No '{body.cas_no}' is already used by another material.")
            skipped += 1
            continue

        consumable_type_id = None
        if consumable_type_name and str(consumable_type_name).strip():
            ctype = db.query(InvConsumableType).filter(
                InvConsumableType.name.ilike(str(consumable_type_name).strip())
            ).first()
            if not ctype:
                errors.append(f"Row {i}: Consumable Type '{consumable_type_name}' not found.")
                skipped += 1
                continue
            consumable_type_id = ctype.id

        seq = _claim_next_seq(db, year)
        data = body.model_dump(exclude={"code", "consumable_type_id"})
        db.add(InvMaterial(code=f"{CODE_PREFIX}/{year}/{seq}", consumable_type_id=consumable_type_id, **data))
        seen_cas[body.cas_no] = i
        created += 1

    db.commit()
    return MaterialUploadResult(created=created, skipped=skipped, errors=errors)
