"""Inventory – Maintenance / Calibration Planner & Schedules (Phase 4).

A schedule row is the due-date record driving the Maintenance Planner
(equipment: MAINTENANCE / CLEANING) and Calibration Planner (instrument:
CALIBRATION). Rows are created manually, via bulk Excel upload, or
auto-generated when a schedule is marked complete.
"""
import calendar
import datetime
from typing import Any, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.orm import Session

from app.dependencies import get_current_user, get_db
from app.models.inventory import (
    InvChecklist,
    InvEquipmentCatalogue,
    InvInstrumentCatalogue,
    InvSchedule,
)
from app.schemas.inventory import (
    ScheduleCompleteRequest,
    ScheduleCreate,
    ScheduleOut,
    ScheduleUpdate,
    ScheduleUploadResult,
)
from app.shared.inv_audit import write_inv_audit

router = APIRouter(prefix="/inventory/schedules", tags=["inventory-schedules"])

SCHEDULE_MONTHS = {"MONTHLY": 1, "QUARTERLY": 3, "HALF_YEARLY": 6, "YEARLY": 12}


def _user_ref(user) -> str:
    return user.username if hasattr(user, "username") else str(user.id)


def _add_months(d: datetime.date, months: int) -> datetime.date:
    month_index = d.month - 1 + months
    year = d.year + month_index // 12
    month = month_index % 12 + 1
    day = min(d.day, calendar.monthrange(year, month)[1])
    return datetime.date(year, month, day)


def _days_label(due_date: datetime.date, done_on: Optional[datetime.date]) -> str:
    today = datetime.date.today()
    ref = done_on or today
    delta = (ref - due_date).days
    if delta > 0:
        return f"{delta} days passed"
    if delta < 0:
        return f"Due in {-delta} days"
    return "Due today"


def _to_out(db: Session, row: InvSchedule) -> dict:
    if row.target_kind == "EQUIPMENT":
        item = db.get(InvEquipmentCatalogue, row.equipment_id) if row.equipment_id else None
        code, kind, current_status = (item.asset_id, "equipment", item.status) if item else (None, None, None)
    else:
        item = db.get(InvInstrumentCatalogue, row.instrument_id) if row.instrument_id else None
        code, kind, current_status = (item.asset_id, "instrument", item.status) if item else (None, None, None)
    return {
        "id": row.id,
        "target_kind": row.target_kind,
        "equipment_id": row.equipment_id,
        "instrument_id": row.instrument_id,
        "equipment_code": code,
        "equipment_type": kind,
        "log_type": row.log_type,
        "checklist_id": row.checklist_id,
        "schedule_type": row.schedule_type,
        "due_date": row.due_date,
        "planned_date": row.planned_date,
        "tolerance_days": row.tolerance_days,
        "alert_limit": row.alert_limit,
        "deviation_limit": row.deviation_limit,
        "done_on": row.done_on,
        "status": row.status,
        "source": row.source,
        "current_status": current_status,
        "days_label": _days_label(row.due_date, row.done_on),
        "created_at": row.created_at,
        "updated_at": row.updated_at,
    }


@router.get("", response_model=list[ScheduleOut])
def list_schedules(
    target_kind: Optional[str] = Query(None),
    log_type: Optional[str] = Query(None),
    schedule_type: Optional[str] = Query(None),
    equipment_id: Optional[int] = Query(None),
    instrument_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    from_date: Optional[datetime.date] = Query(None),
    to_date: Optional[datetime.date] = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(200, le=1000),
    db: Session = Depends(get_db),
    _: Any = Depends(get_current_user),
):
    q = db.query(InvSchedule)
    if target_kind:
        q = q.filter(InvSchedule.target_kind == target_kind)
    if log_type:
        q = q.filter(InvSchedule.log_type == log_type)
    if schedule_type:
        q = q.filter(InvSchedule.schedule_type == schedule_type)
    if equipment_id is not None:
        q = q.filter(InvSchedule.equipment_id == equipment_id)
    if instrument_id is not None:
        q = q.filter(InvSchedule.instrument_id == instrument_id)
    if status:
        q = q.filter(InvSchedule.status == status)
    if from_date:
        q = q.filter(InvSchedule.due_date >= from_date)
    if to_date:
        q = q.filter(InvSchedule.due_date <= to_date)
    rows = q.order_by(InvSchedule.due_date).offset(skip).limit(limit).all()
    return [_to_out(db, r) for r in rows]


def _validate_target(db: Session, body_or_row) -> str:
    if bool(body_or_row.equipment_id) == bool(body_or_row.instrument_id):
        raise HTTPException(400, "Provide exactly one of equipment_id or instrument_id.")
    if body_or_row.equipment_id:
        if not db.get(InvEquipmentCatalogue, body_or_row.equipment_id):
            raise HTTPException(404, "Equipment not found.")
        return "EQUIPMENT"
    if not db.get(InvInstrumentCatalogue, body_or_row.instrument_id):
        raise HTTPException(404, "Instrument not found.")
    return "INSTRUMENT"


@router.post("", response_model=ScheduleOut, status_code=201)
def create_schedule(
    body: ScheduleCreate,
    db: Session = Depends(get_db),
    current_user: Any = Depends(get_current_user),
):
    target_kind = _validate_target(db, body)
    if body.schedule_type not in SCHEDULE_MONTHS:
        raise HTTPException(400, f"schedule_type must be one of {list(SCHEDULE_MONTHS)}.")
    if body.checklist_id:
        cl = db.get(InvChecklist, body.checklist_id)
        if not cl or cl.status != "APPROVED":
            raise HTTPException(409, "Only APPROVED checklists can be scheduled.")
    row = InvSchedule(**body.model_dump(), target_kind=target_kind, source="MANUAL", created_by=_user_ref(current_user))
    db.add(row)
    db.commit()
    db.refresh(row)
    return _to_out(db, row)


@router.patch("/{schedule_id}", response_model=ScheduleOut)
def update_schedule(
    schedule_id: int,
    body: ScheduleUpdate,
    db: Session = Depends(get_db),
    _: Any = Depends(get_current_user),
):
    row = db.get(InvSchedule, schedule_id)
    if not row:
        raise HTTPException(404, "Schedule not found.")
    for k, v in body.model_dump(exclude_none=True).items():
        setattr(row, k, v)
    db.commit()
    db.refresh(row)
    return _to_out(db, row)


@router.delete("/{schedule_id}", status_code=204)
def delete_schedule(
    schedule_id: int,
    db: Session = Depends(get_db),
    _: Any = Depends(get_current_user),
):
    row = db.get(InvSchedule, schedule_id)
    if not row:
        raise HTTPException(404, "Schedule not found.")
    db.delete(row)
    db.commit()


@router.post("/{schedule_id}/complete", response_model=ScheduleOut)
def complete_schedule(
    schedule_id: int,
    body: ScheduleCompleteRequest,
    db: Session = Depends(get_db),
    current_user: Any = Depends(get_current_user),
):
    row = db.get(InvSchedule, schedule_id)
    if not row:
        raise HTTPException(404, "Schedule not found.")
    if row.status == "DONE":
        raise HTTPException(409, "Schedule already marked done.")
    row.done_on = body.done_on
    row.status = "DONE"
    write_inv_audit(
        db,
        event_type="SCHEDULE_COMPLETED",
        entity_type="inv_schedule",
        entity_id=row.id,
        entity_ref=f"{row.log_type} #{row.equipment_id or row.instrument_id}",
        performed_by=_user_ref(current_user),
        new_value=str(body.done_on),
    )
    next_row = None
    if body.generate_next:
        months = SCHEDULE_MONTHS.get(row.schedule_type, 1)
        next_due = _add_months(row.due_date, months)
        next_row = InvSchedule(
            target_kind=row.target_kind, equipment_id=row.equipment_id, instrument_id=row.instrument_id,
            log_type=row.log_type, checklist_id=row.checklist_id, schedule_type=row.schedule_type,
            due_date=next_due, tolerance_days=row.tolerance_days, alert_limit=row.alert_limit,
            deviation_limit=row.deviation_limit, status="DUE", source="AUTO_GENERATED",
            created_by=_user_ref(current_user),
        )
        db.add(next_row)
    db.commit()
    db.refresh(row)
    return _to_out(db, row)


# ── Bulk Excel upload ──────────────────────────────────────────────────────────
EXPECTED_HEADERS = ["Equipment/Instrument Code", "Schedule Type", "Due Date"]


@router.post("/upload", response_model=ScheduleUploadResult)
def upload_schedules(
    target_kind: str = Query(..., description="EQUIPMENT or INSTRUMENT"),
    log_type: str = Query(..., description="MAINTENANCE, CLEANING or CALIBRATION"),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: Any = Depends(get_current_user),
):
    from openpyxl import load_workbook

    if target_kind not in ("EQUIPMENT", "INSTRUMENT"):
        raise HTTPException(400, "target_kind must be EQUIPMENT or INSTRUMENT.")

    try:
        wb = load_workbook(file.file, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(400, f"Could not read Excel file: {exc}")
    ws = wb.active

    created, skipped, errors = 0, 0, []
    Model = InvEquipmentCatalogue if target_kind == "EQUIPMENT" else InvInstrumentCatalogue
    id_field = "equipment_id" if target_kind == "EQUIPMENT" else "instrument_id"

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(c is None for c in row):
            continue
        code, schedule_type, due_date = (row + (None, None, None))[:3]
        if not code or not schedule_type or not due_date:
            errors.append(f"Row {i}: missing required value(s).")
            skipped += 1
            continue
        item = db.query(Model).filter_by(asset_id=str(code).strip()).first()
        if not item:
            errors.append(f"Row {i}: code '{code}' not found.")
            skipped += 1
            continue
        schedule_type = str(schedule_type).strip().upper().replace(" ", "_")
        if schedule_type not in SCHEDULE_MONTHS:
            errors.append(f"Row {i}: invalid schedule type '{schedule_type}'.")
            skipped += 1
            continue
        if isinstance(due_date, datetime.datetime):
            due_date = due_date.date()
        elif isinstance(due_date, str):
            try:
                due_date = datetime.date.fromisoformat(due_date)
            except ValueError:
                errors.append(f"Row {i}: invalid date '{due_date}'.")
                skipped += 1
                continue
        db.add(InvSchedule(
            target_kind=target_kind, log_type=log_type, schedule_type=schedule_type,
            due_date=due_date, status="DUE", source="EXCEL_UPLOAD",
            created_by=_user_ref(current_user), **{id_field: item.id},
        ))
        created += 1

    write_inv_audit(
        db,
        event_type="SCHEDULE_BULK_UPLOAD",
        entity_type="inv_schedule",
        performed_by=_user_ref(current_user),
        details=f"{target_kind}/{log_type}: {created} created, {skipped} skipped",
    )
    db.commit()
    return ScheduleUploadResult(created=created, skipped=skipped, errors=errors)
