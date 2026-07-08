"""Seed: inv_materials from the AB-ADC Consumable Inventory workbook.

Upserts by `code` (idempotent). Resolves Consumable Type name -> consumable_type_id
via inv_consumable_types (run seed_consumable_types.py first).

Usage:
    python seeds/seed_materials_from_excel.py [path\\to\\workbook.xlsx]

Defaults to the reformatted workbook in the current user's Downloads folder.
"""
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import load_workbook

from app.database import SessionLocal
# Import every model module so all FK targets (departments, etc.) are registered.
from app.models import (  # noqa: F401
    adc, admin, experiment, inventory, master_data, master_data_admin,
    notebook, project, settings, workflow_template,
)
from app.models.inventory import InvConsumableType, InvMaterial

DEFAULT_XLSX = os.path.join(
    os.path.expanduser("~"), "Downloads",
    "AB-ADC_Consumable_Inventory_Reformatted.xlsx",
)
SHEET = "Consumable Inventory"

# Excel column order (1-based): Code, Name, Material Type, CAS NO, Molecular
# Formula, Mol. Weight, Storage Condition, Hazard Class, Consumable Type,
# Description, Source Sheet
COL = dict(code=1, name=2, material_type=3, cas_no=4, molecular_formula=5,
           mol_weight=6, storage_condition=7, hazard_class=8,
           consumable_type=9, description=10)


def _clean(v):
    if v is None:
        return None
    s = str(v).strip()
    if s == "" or s.upper() == "N/A":
        return None
    return s


def _num(v):
    s = _clean(v)
    if s is None:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def run(path=DEFAULT_XLSX):
    if not os.path.isfile(path):
        raise SystemExit(f"Workbook not found: {path}")

    wb = load_workbook(path, data_only=True)
    ws = wb[SHEET]

    db = SessionLocal()
    try:
        ct_by_name = {c.name.lower(): c.id for c in db.query(InvConsumableType)}
        inserted = updated = 0
        for r in range(2, ws.max_row + 1):
            code = _clean(ws.cell(r, COL["code"]).value)
            name = _clean(ws.cell(r, COL["name"]).value)
            if not code or not name:
                continue

            ct_name = _clean(ws.cell(r, COL["consumable_type"]).value)
            ct_id = ct_by_name.get(ct_name.lower()) if ct_name else None

            fields = dict(
                name=name,
                material_type=_clean(ws.cell(r, COL["material_type"]).value),
                cas_no=_clean(ws.cell(r, COL["cas_no"]).value),
                molecular_formula=_clean(ws.cell(r, COL["molecular_formula"]).value),
                mol_weight=_num(ws.cell(r, COL["mol_weight"]).value),
                storage_condition=_clean(ws.cell(r, COL["storage_condition"]).value),
                hazard_class=_clean(ws.cell(r, COL["hazard_class"]).value),
                description=_clean(ws.cell(r, COL["description"]).value),
                consumable_type_id=ct_id,
                is_active=True,
            )

            row = db.query(InvMaterial).filter_by(code=code).first()
            if row:
                for k, v in fields.items():
                    setattr(row, k, v)
                updated += 1
            else:
                db.add(InvMaterial(code=code, **fields))
                inserted += 1

        db.commit()
        total = db.query(InvMaterial).count()
        print(f"seed_materials_from_excel: {inserted} inserted, {updated} updated. "
              f"inv_materials total = {total}.")
    finally:
        db.close()


if __name__ == "__main__":
    run(sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX)
