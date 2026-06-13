"""
Generates Inventory Master — Database Schema Reference Excel file.
Run: python generate_inventory_schema.py
Output: Inventory_Master_DB_Schema.xlsx (same directory)
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ─── Colour palette ───────────────────────────────────────────────────────────

TEAL_DARK   = "0F766E"   # header bg
TEAL_MID    = "0D9488"   # section title
TEAL_LIGHT  = "CCFBF1"   # column header row
STONE_50    = "FAFAF9"   # even row
WHITE       = "FFFFFF"
AMBER_LIGHT = "FEF3C7"   # FK highlight
RED_LIGHT   = "FEE2E2"   # NOT NULL highlight
GREEN_LIGHT = "DCFCE7"   # PK highlight
GRAY_200    = "E7E5E4"   # border

# ─── Helper styles ────────────────────────────────────────────────────────────

def thin_border(color=GRAY_200):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def header_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center_align(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left_align(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

# ─── Full schema definition ───────────────────────────────────────────────────
# Each table: { name, description, group, columns: [ { col, type, nullable, default, constraints, notes } ] }

SCHEMA = [

    # ── 1. inv_materials ──────────────────────────────────────────────────────
    {
        "name": "inv_materials",
        "description": "Core materials master — chemicals, reagents, solvents, APIs, etc.",
        "group": "Materials",
        "columns": [
            {"col": "id",                "type": "INTEGER",        "nullable": "NO",  "default": "auto-increment", "constraints": "PRIMARY KEY, INDEX",        "notes": ""},
            {"col": "code",              "type": "VARCHAR(50)",    "nullable": "NO",  "default": "",               "constraints": "UNIQUE, NOT NULL, INDEX",    "notes": "e.g. MAT-0001"},
            {"col": "name",              "type": "VARCHAR(255)",   "nullable": "NO",  "default": "",               "constraints": "NOT NULL",                  "notes": "Full material name"},
            {"col": "material_type",     "type": "VARCHAR(100)",   "nullable": "YES", "default": "",               "constraints": "",                          "notes": "Chemical / Reagent / Solvent / Standard / API / etc."},
            {"col": "cas_no",            "type": "VARCHAR(50)",    "nullable": "YES", "default": "",               "constraints": "",                          "notes": "CAS registry number"},
            {"col": "molecular_formula", "type": "VARCHAR(100)",   "nullable": "YES", "default": "",               "constraints": "",                          "notes": "e.g. C6H12O6"},
            {"col": "mol_weight",        "type": "NUMERIC(12,4)",  "nullable": "YES", "default": "",               "constraints": "",                          "notes": "g/mol"},
            {"col": "storage_condition", "type": "VARCHAR(200)",   "nullable": "YES", "default": "",               "constraints": "",                          "notes": "e.g. 2–8°C, Keep dry"},
            {"col": "hazard_class",      "type": "VARCHAR(100)",   "nullable": "YES", "default": "",               "constraints": "",                          "notes": "e.g. Flammable, Corrosive"},
            {"col": "description",       "type": "TEXT",           "nullable": "YES", "default": "",               "constraints": "",                          "notes": ""},
            {"col": "is_active",         "type": "BOOLEAN",        "nullable": "NO",  "default": "TRUE",           "constraints": "NOT NULL",                  "notes": "Soft-delete flag"},
        ],
    },

    # ── 2. inv_material_chemical_props ────────────────────────────────────────
    {
        "name": "inv_material_chemical_props",
        "description": "Optional chemical/physical properties for a material (1:1 with inv_materials).",
        "group": "Materials",
        "columns": [
            {"col": "id",          "type": "INTEGER",       "nullable": "NO",  "default": "auto-increment", "constraints": "PRIMARY KEY, INDEX",                             "notes": ""},
            {"col": "material_id", "type": "INTEGER",       "nullable": "NO",  "default": "",               "constraints": "FK → inv_materials.id (CASCADE), UNIQUE, NOT NULL","notes": "1:1 link to material"},
            {"col": "purity_pct",  "type": "NUMERIC(6,2)",  "nullable": "YES", "default": "",               "constraints": "",                                               "notes": "%"},
            {"col": "grade",       "type": "VARCHAR(100)",  "nullable": "YES", "default": "",               "constraints": "",                                               "notes": "e.g. HPLC, ACS"},
            {"col": "appearance",  "type": "VARCHAR(200)",  "nullable": "YES", "default": "",               "constraints": "",                                               "notes": "e.g. White crystalline powder"},
            {"col": "solubility",  "type": "VARCHAR(200)",  "nullable": "YES", "default": "",               "constraints": "",                                               "notes": ""},
            {"col": "boiling_pt",  "type": "VARCHAR(50)",   "nullable": "YES", "default": "",               "constraints": "",                                               "notes": "°C (stored as string for ranges)"},
            {"col": "melting_pt",  "type": "VARCHAR(50)",   "nullable": "YES", "default": "",               "constraints": "",                                               "notes": "°C"},
            {"col": "flash_pt",    "type": "VARCHAR(50)",   "nullable": "YES", "default": "",               "constraints": "",                                               "notes": "°C"},
            {"col": "density",     "type": "NUMERIC(8,4)",  "nullable": "YES", "default": "",               "constraints": "",                                               "notes": "g/mL"},
            {"col": "ph_range",    "type": "VARCHAR(50)",   "nullable": "YES", "default": "",               "constraints": "",                                               "notes": "e.g. 6.5–7.5"},
        ],
    },

    # ── 3. inv_material_formulation_props ─────────────────────────────────────
    {
        "name": "inv_material_formulation_props",
        "description": "Optional formulation role/properties for a material (1:1 with inv_materials).",
        "group": "Materials",
        "columns": [
            {"col": "id",                  "type": "INTEGER",      "nullable": "NO",  "default": "auto-increment", "constraints": "PRIMARY KEY, INDEX",                              "notes": ""},
            {"col": "material_id",         "type": "INTEGER",      "nullable": "NO",  "default": "",               "constraints": "FK → inv_materials.id (CASCADE), UNIQUE, NOT NULL","notes": "1:1 link to material"},
            {"col": "role",                "type": "VARCHAR(100)", "nullable": "YES", "default": "",               "constraints": "",                                                "notes": "e.g. Excipient, Active, Buffer"},
            {"col": "concentration",       "type": "VARCHAR(100)", "nullable": "YES", "default": "",               "constraints": "",                                                "notes": "Value + unit string"},
            {"col": "units",               "type": "VARCHAR(50)",  "nullable": "YES", "default": "",               "constraints": "",                                                "notes": "e.g. mg/mL, %w/v"},
            {"col": "function",            "type": "TEXT",         "nullable": "YES", "default": "",               "constraints": "",                                                "notes": ""},
            {"col": "compatibility_notes", "type": "TEXT",         "nullable": "YES", "default": "",               "constraints": "",                                                "notes": ""},
        ],
    },

    # ── 4. inv_manufacturers ──────────────────────────────────────────────────
    {
        "name": "inv_manufacturers",
        "description": "Supplier / manufacturer master list.",
        "group": "Manufacturers",
        "columns": [
            {"col": "id",             "type": "INTEGER",      "nullable": "NO",  "default": "auto-increment", "constraints": "PRIMARY KEY, INDEX",     "notes": ""},
            {"col": "code",           "type": "VARCHAR(50)",  "nullable": "NO",  "default": "",               "constraints": "UNIQUE, NOT NULL, INDEX", "notes": "e.g. MFR-001"},
            {"col": "name",           "type": "VARCHAR(255)", "nullable": "NO",  "default": "",               "constraints": "NOT NULL",               "notes": "Company/supplier name"},
            {"col": "country",        "type": "VARCHAR(100)", "nullable": "YES", "default": "",               "constraints": "",                       "notes": ""},
            {"col": "contact_person", "type": "VARCHAR(200)", "nullable": "YES", "default": "",               "constraints": "",                       "notes": ""},
            {"col": "email",          "type": "VARCHAR(200)", "nullable": "YES", "default": "",               "constraints": "",                       "notes": ""},
            {"col": "phone",          "type": "VARCHAR(50)",  "nullable": "YES", "default": "",               "constraints": "",                       "notes": ""},
            {"col": "website",        "type": "VARCHAR(300)", "nullable": "YES", "default": "",               "constraints": "",                       "notes": ""},
            {"col": "address",        "type": "TEXT",         "nullable": "YES", "default": "",               "constraints": "",                       "notes": ""},
            {"col": "is_active",      "type": "BOOLEAN",      "nullable": "NO",  "default": "TRUE",           "constraints": "NOT NULL",               "notes": "Soft-delete flag"},
        ],
    },

    # ── 5. inv_manufacturer_mapping ───────────────────────────────────────────
    {
        "name": "inv_manufacturer_mapping",
        "description": "Links a material to one or more manufacturers with sourcing details.",
        "group": "Manufacturers",
        "columns": [
            {"col": "id",              "type": "INTEGER",      "nullable": "NO",  "default": "auto-increment", "constraints": "PRIMARY KEY, INDEX",                    "notes": ""},
            {"col": "material_id",     "type": "INTEGER",      "nullable": "NO",  "default": "",               "constraints": "FK → inv_materials.id (CASCADE), NOT NULL","notes": ""},
            {"col": "manufacturer_id", "type": "INTEGER",      "nullable": "NO",  "default": "",               "constraints": "FK → inv_manufacturers.id (CASCADE), NOT NULL","notes": ""},
            {"col": "catalogue_no",    "type": "VARCHAR(100)", "nullable": "YES", "default": "",               "constraints": "",                                      "notes": "Manufacturer catalogue number"},
            {"col": "technical_grade", "type": "VARCHAR(100)", "nullable": "YES", "default": "",               "constraints": "",                                      "notes": ""},
            {"col": "lead_time_days",  "type": "INTEGER",      "nullable": "YES", "default": "",               "constraints": "",                                      "notes": "Typical delivery days"},
            {"col": "min_order_qty",   "type": "NUMERIC(10,3)","nullable": "YES", "default": "",               "constraints": "",                                      "notes": "Minimum order quantity"},
        ],
    },

    # ── 6. inv_batches ────────────────────────────────────────────────────────
    {
        "name": "inv_batches",
        "description": "Inventory batch records — received stock with qty tracking.",
        "group": "Batches",
        "columns": [
            {"col": "id",              "type": "INTEGER",           "nullable": "NO",  "default": "auto-increment",    "constraints": "PRIMARY KEY, INDEX",              "notes": ""},
            {"col": "batch_no",        "type": "VARCHAR(100)",      "nullable": "NO",  "default": "",                  "constraints": "UNIQUE, NOT NULL, INDEX",          "notes": "e.g. BATCH-2025-001"},
            {"col": "material_id",     "type": "INTEGER",           "nullable": "NO",  "default": "",                  "constraints": "FK → inv_materials.id, NOT NULL",  "notes": ""},
            {"col": "manufacturer_id", "type": "INTEGER",           "nullable": "YES", "default": "",                  "constraints": "FK → inv_manufacturers.id",        "notes": ""},
            {"col": "qty_received",    "type": "NUMERIC(12,4)",     "nullable": "NO",  "default": "",                  "constraints": "NOT NULL",                        "notes": "Original received quantity"},
            {"col": "qty_available",   "type": "NUMERIC(12,4)",     "nullable": "NO",  "default": "",                  "constraints": "NOT NULL",                        "notes": "Current remaining quantity"},
            {"col": "unit",            "type": "VARCHAR(20)",       "nullable": "NO",  "default": "'g'",               "constraints": "NOT NULL",                        "notes": "g / mg / mL / L / units"},
            {"col": "location",        "type": "VARCHAR(200)",      "nullable": "YES", "default": "",                  "constraints": "",                                "notes": "Storage location / room / fridge"},
            {"col": "mfg_date",        "type": "DATE",              "nullable": "YES", "default": "",                  "constraints": "",                                "notes": "Manufacture date"},
            {"col": "expiry_date",     "type": "DATE",              "nullable": "YES", "default": "",                  "constraints": "",                                "notes": "Expiry / use-by date"},
            {"col": "retest_date",     "type": "DATE",              "nullable": "YES", "default": "",                  "constraints": "",                                "notes": "Re-test / review date"},
            {"col": "invoice_no",      "type": "VARCHAR(100)",      "nullable": "YES", "default": "",                  "constraints": "",                                "notes": ""},
            {"col": "po_no",           "type": "VARCHAR(100)",      "nullable": "YES", "default": "",                  "constraints": "",                                "notes": "Purchase order number"},
            {"col": "remarks",         "type": "TEXT",              "nullable": "YES", "default": "",                  "constraints": "",                                "notes": ""},
            {"col": "status",          "type": "VARCHAR(30)",       "nullable": "NO",  "default": "'AVAILABLE'",       "constraints": "NOT NULL",                        "notes": "AVAILABLE | PARTIALLY_CONSUMED | CONSUMED | EXPIRED | QUARANTINE"},
            {"col": "category",        "type": "VARCHAR(20)",       "nullable": "NO",  "default": "'available'",       "constraints": "NOT NULL",                        "notes": "available | non_available | historic"},
            {"col": "received_by",     "type": "VARCHAR(200)",      "nullable": "YES", "default": "",                  "constraints": "",                                "notes": "Username who received the batch"},
            {"col": "received_at",     "type": "TIMESTAMPTZ",       "nullable": "YES", "default": "now()",             "constraints": "SERVER DEFAULT",                  "notes": "Auto-set on insert"},
            {"col": "is_active",       "type": "BOOLEAN",           "nullable": "NO",  "default": "TRUE",              "constraints": "NOT NULL",                        "notes": "Soft-delete flag"},
        ],
    },

    # ── 7. inv_batch_events ───────────────────────────────────────────────────
    {
        "name": "inv_batch_events",
        "description": "Audit log of all quantity movements for a batch.",
        "group": "Batches",
        "columns": [
            {"col": "id",           "type": "INTEGER",       "nullable": "NO",  "default": "auto-increment", "constraints": "PRIMARY KEY, INDEX",              "notes": ""},
            {"col": "batch_id",     "type": "INTEGER",       "nullable": "NO",  "default": "",               "constraints": "FK → inv_batches.id (CASCADE), NOT NULL","notes": ""},
            {"col": "event_type",   "type": "VARCHAR(50)",   "nullable": "NO",  "default": "",               "constraints": "NOT NULL",                        "notes": "RECEIVED | ISSUED | STOCK_ALLOCATION | LABEL_GENERATED | ADJUSTMENT | DISPOSAL"},
            {"col": "qty",          "type": "NUMERIC(12,4)", "nullable": "YES", "default": "",               "constraints": "",                                "notes": "Quantity involved in event"},
            {"col": "ref_no",       "type": "VARCHAR(100)",  "nullable": "YES", "default": "",               "constraints": "",                                "notes": "Reference number (experiment / request)"},
            {"col": "module",       "type": "VARCHAR(100)",  "nullable": "YES", "default": "",               "constraints": "",                                "notes": "ADC / QC / R&D"},
            {"col": "issued_to",    "type": "VARCHAR(200)",  "nullable": "YES", "default": "",               "constraints": "",                                "notes": "Recipient username"},
            {"col": "purpose",      "type": "TEXT",          "nullable": "YES", "default": "",               "constraints": "",                                "notes": ""},
            {"col": "project_code", "type": "VARCHAR(100)",  "nullable": "YES", "default": "",               "constraints": "",                                "notes": ""},
            {"col": "performed_by", "type": "VARCHAR(200)",  "nullable": "YES", "default": "",               "constraints": "",                                "notes": ""},
            {"col": "performed_at", "type": "TIMESTAMPTZ",   "nullable": "YES", "default": "now()",          "constraints": "SERVER DEFAULT",                  "notes": ""},
            {"col": "remarks",      "type": "TEXT",          "nullable": "YES", "default": "",               "constraints": "",                                "notes": ""},
        ],
    },

    # ── 8. inv_batch_verifications ────────────────────────────────────────────
    {
        "name": "inv_batch_verifications",
        "description": "Batch quality verification requests (QC sign-off).",
        "group": "Batches",
        "columns": [
            {"col": "id",           "type": "INTEGER",     "nullable": "NO",  "default": "auto-increment", "constraints": "PRIMARY KEY, INDEX",              "notes": ""},
            {"col": "request_no",   "type": "VARCHAR(100)","nullable": "NO",  "default": "",               "constraints": "UNIQUE, NOT NULL, INDEX",          "notes": "e.g. BVR-2025-001"},
            {"col": "batch_id",     "type": "INTEGER",     "nullable": "NO",  "default": "",               "constraints": "FK → inv_batches.id (CASCADE), NOT NULL","notes": ""},
            {"col": "requested_by", "type": "VARCHAR(200)","nullable": "YES", "default": "",               "constraints": "",                                "notes": ""},
            {"col": "requested_at", "type": "TIMESTAMPTZ", "nullable": "YES", "default": "now()",          "constraints": "SERVER DEFAULT",                  "notes": ""},
            {"col": "verified_by",  "type": "VARCHAR(200)","nullable": "YES", "default": "",               "constraints": "",                                "notes": ""},
            {"col": "verified_at",  "type": "TIMESTAMPTZ", "nullable": "YES", "default": "",               "constraints": "",                                "notes": ""},
            {"col": "status",       "type": "VARCHAR(20)", "nullable": "NO",  "default": "'PENDING'",      "constraints": "NOT NULL",                        "notes": "PENDING | VERIFIED | REJECTED"},
            {"col": "remarks",      "type": "TEXT",        "nullable": "YES", "default": "",               "constraints": "",                                "notes": ""},
        ],
    },

    # ── 9. inv_stock_requests ─────────────────────────────────────────────────
    {
        "name": "inv_stock_requests",
        "description": "Material stock requisition requests with approval workflow.",
        "group": "Stock Requests",
        "columns": [
            {"col": "id",               "type": "INTEGER",       "nullable": "NO",  "default": "auto-increment", "constraints": "PRIMARY KEY, INDEX",              "notes": ""},
            {"col": "request_no",       "type": "VARCHAR(100)",  "nullable": "NO",  "default": "",               "constraints": "UNIQUE, NOT NULL, INDEX",          "notes": "e.g. SR-2025-001"},
            {"col": "material_id",      "type": "INTEGER",       "nullable": "NO",  "default": "",               "constraints": "FK → inv_materials.id, NOT NULL",  "notes": ""},
            {"col": "qty_required",     "type": "NUMERIC(12,4)", "nullable": "NO",  "default": "",               "constraints": "NOT NULL",                        "notes": ""},
            {"col": "unit",             "type": "VARCHAR(20)",   "nullable": "NO",  "default": "'g'",            "constraints": "NOT NULL",                        "notes": ""},
            {"col": "required_by_date", "type": "DATE",          "nullable": "YES", "default": "",               "constraints": "",                                "notes": "Deadline for fulfilment"},
            {"col": "criticality",      "type": "VARCHAR(20)",   "nullable": "NO",  "default": "'MEDIUM'",       "constraints": "NOT NULL",                        "notes": "LOW | MEDIUM | HIGH | CRITICAL"},
            {"col": "purpose",          "type": "TEXT",          "nullable": "YES", "default": "",               "constraints": "",                                "notes": ""},
            {"col": "requested_by",     "type": "VARCHAR(200)",  "nullable": "YES", "default": "",               "constraints": "",                                "notes": ""},
            {"col": "requested_at",     "type": "TIMESTAMPTZ",   "nullable": "YES", "default": "now()",          "constraints": "SERVER DEFAULT",                  "notes": ""},
            {"col": "approved_by",      "type": "VARCHAR(200)",  "nullable": "YES", "default": "",               "constraints": "",                                "notes": ""},
            {"col": "approved_at",      "type": "TIMESTAMPTZ",   "nullable": "YES", "default": "",               "constraints": "",                                "notes": ""},
            {"col": "status",           "type": "VARCHAR(20)",   "nullable": "NO",  "default": "'PENDING'",      "constraints": "NOT NULL",                        "notes": "PENDING | APPROVED | REJECTED | FULFILLED | CANCELLED"},
            {"col": "remarks",          "type": "TEXT",          "nullable": "YES", "default": "",               "constraints": "",                                "notes": ""},
        ],
    },

    # ── 10. inv_stock_request_events ──────────────────────────────────────────
    {
        "name": "inv_stock_request_events",
        "description": "Lifecycle event log for stock requests.",
        "group": "Stock Requests",
        "columns": [
            {"col": "id",           "type": "INTEGER",     "nullable": "NO",  "default": "auto-increment", "constraints": "PRIMARY KEY, INDEX",                   "notes": ""},
            {"col": "request_id",   "type": "INTEGER",     "nullable": "NO",  "default": "",               "constraints": "FK → inv_stock_requests.id (CASCADE), NOT NULL","notes": ""},
            {"col": "event_type",   "type": "VARCHAR(50)", "nullable": "NO",  "default": "",               "constraints": "NOT NULL",                             "notes": "SUBMITTED | APPROVED | REJECTED | FULFILLED | CANCELLED | UPDATED"},
            {"col": "performed_by", "type": "VARCHAR(200)","nullable": "YES", "default": "",               "constraints": "",                                     "notes": ""},
            {"col": "performed_at", "type": "TIMESTAMPTZ", "nullable": "YES", "default": "now()",          "constraints": "SERVER DEFAULT",                       "notes": ""},
            {"col": "remarks",      "type": "TEXT",        "nullable": "YES", "default": "",               "constraints": "",                                     "notes": ""},
        ],
    },

    # ── 11. inv_equipment_types ───────────────────────────────────────────────
    {
        "name": "inv_equipment_types",
        "description": "Equipment type master (e.g. Centrifuge, Autoclave).",
        "group": "Equipment Master",
        "columns": [
            {"col": "id",          "type": "INTEGER",     "nullable": "NO",  "default": "auto-increment", "constraints": "PRIMARY KEY, INDEX",     "notes": ""},
            {"col": "code",        "type": "VARCHAR(50)", "nullable": "NO",  "default": "",               "constraints": "UNIQUE, NOT NULL, INDEX", "notes": ""},
            {"col": "name",        "type": "VARCHAR(255)","nullable": "NO",  "default": "",               "constraints": "NOT NULL",               "notes": ""},
            {"col": "description", "type": "TEXT",        "nullable": "YES", "default": "",               "constraints": "",                       "notes": ""},
            {"col": "is_active",   "type": "BOOLEAN",     "nullable": "NO",  "default": "TRUE",           "constraints": "NOT NULL",               "notes": ""},
        ],
    },

    # ── 12. inv_instrument_types ──────────────────────────────────────────────
    {
        "name": "inv_instrument_types",
        "description": "Instrument type master (e.g. HPLC, UV-Vis, Mass Spec).",
        "group": "Equipment Master",
        "columns": [
            {"col": "id",          "type": "INTEGER",     "nullable": "NO",  "default": "auto-increment", "constraints": "PRIMARY KEY, INDEX",     "notes": ""},
            {"col": "code",        "type": "VARCHAR(50)", "nullable": "NO",  "default": "",               "constraints": "UNIQUE, NOT NULL, INDEX", "notes": ""},
            {"col": "name",        "type": "VARCHAR(255)","nullable": "NO",  "default": "",               "constraints": "NOT NULL",               "notes": ""},
            {"col": "description", "type": "TEXT",        "nullable": "YES", "default": "",               "constraints": "",                       "notes": ""},
            {"col": "is_active",   "type": "BOOLEAN",     "nullable": "NO",  "default": "TRUE",           "constraints": "NOT NULL",               "notes": ""},
        ],
    },

    # ── 13. inv_column_types ──────────────────────────────────────────────────
    {
        "name": "inv_column_types",
        "description": "Chromatography column type master with physical specs.",
        "group": "Equipment Master",
        "columns": [
            {"col": "id",                 "type": "INTEGER",     "nullable": "NO",  "default": "auto-increment", "constraints": "PRIMARY KEY, INDEX",     "notes": ""},
            {"col": "code",               "type": "VARCHAR(50)", "nullable": "NO",  "default": "",               "constraints": "UNIQUE, NOT NULL, INDEX", "notes": ""},
            {"col": "name",               "type": "VARCHAR(255)","nullable": "NO",  "default": "",               "constraints": "NOT NULL",               "notes": ""},
            {"col": "description",        "type": "TEXT",        "nullable": "YES", "default": "",               "constraints": "",                       "notes": ""},
            {"col": "length_mm",          "type": "NUMERIC(8,2)","nullable": "YES", "default": "",               "constraints": "",                       "notes": "Column length in mm"},
            {"col": "particle_size_um",   "type": "NUMERIC(8,2)","nullable": "YES", "default": "",               "constraints": "",                       "notes": "Particle size in µm"},
            {"col": "pore_size_angstrom", "type": "NUMERIC(8,2)","nullable": "YES", "default": "",               "constraints": "",                       "notes": "Pore size in Å"},
            {"col": "is_active",          "type": "BOOLEAN",     "nullable": "NO",  "default": "TRUE",           "constraints": "NOT NULL",               "notes": ""},
        ],
    },

    # ── 14. inv_equipment_catalogue ───────────────────────────────────────────
    {
        "name": "inv_equipment_catalogue",
        "description": "Physical equipment asset register with maintenance tracking.",
        "group": "Catalogues",
        "columns": [
            {"col": "id",                    "type": "INTEGER",     "nullable": "NO",  "default": "auto-increment", "constraints": "PRIMARY KEY, INDEX",                    "notes": ""},
            {"col": "asset_id",              "type": "VARCHAR(100)","nullable": "NO",  "default": "",               "constraints": "UNIQUE, NOT NULL, INDEX",               "notes": "e.g. EQP-001"},
            {"col": "name",                  "type": "VARCHAR(255)","nullable": "NO",  "default": "",               "constraints": "NOT NULL",                             "notes": ""},
            {"col": "equipment_type_id",     "type": "INTEGER",     "nullable": "YES", "default": "",               "constraints": "FK → inv_equipment_types.id",          "notes": ""},
            {"col": "serial_no",             "type": "VARCHAR(100)","nullable": "YES", "default": "",               "constraints": "",                                     "notes": ""},
            {"col": "manufacturer",          "type": "VARCHAR(200)","nullable": "YES", "default": "",               "constraints": "",                                     "notes": "Brand/OEM name"},
            {"col": "model",                 "type": "VARCHAR(200)","nullable": "YES", "default": "",               "constraints": "",                                     "notes": ""},
            {"col": "location",              "type": "VARCHAR(200)","nullable": "YES", "default": "",               "constraints": "",                                     "notes": ""},
            {"col": "purchase_date",         "type": "DATE",        "nullable": "YES", "default": "",               "constraints": "",                                     "notes": ""},
            {"col": "last_maintenance_date", "type": "DATE",        "nullable": "YES", "default": "",               "constraints": "",                                     "notes": ""},
            {"col": "maintenance_due_date",  "type": "DATE",        "nullable": "YES", "default": "",               "constraints": "",                                     "notes": ""},
            {"col": "maintenance_status",    "type": "VARCHAR(20)", "nullable": "NO",  "default": "'OK'",           "constraints": "NOT NULL",                             "notes": "OK | DUE | OVERDUE"},
            {"col": "status",                "type": "VARCHAR(30)", "nullable": "NO",  "default": "'ACTIVE'",       "constraints": "NOT NULL",                             "notes": "ACTIVE | INACTIVE | UNDER_MAINTENANCE | DECOMMISSIONED"},
            {"col": "is_active",             "type": "BOOLEAN",     "nullable": "NO",  "default": "TRUE",           "constraints": "NOT NULL",                             "notes": ""},
        ],
    },

    # ── 15. inv_instrument_catalogue ──────────────────────────────────────────
    {
        "name": "inv_instrument_catalogue",
        "description": "Analytical instrument register with calibration tracking.",
        "group": "Catalogues",
        "columns": [
            {"col": "id",                    "type": "INTEGER",     "nullable": "NO",  "default": "auto-increment", "constraints": "PRIMARY KEY, INDEX",                   "notes": ""},
            {"col": "asset_id",              "type": "VARCHAR(100)","nullable": "NO",  "default": "",               "constraints": "UNIQUE, NOT NULL, INDEX",              "notes": "e.g. INST-001"},
            {"col": "name",                  "type": "VARCHAR(255)","nullable": "NO",  "default": "",               "constraints": "NOT NULL",                            "notes": ""},
            {"col": "instrument_type_id",    "type": "INTEGER",     "nullable": "YES", "default": "",               "constraints": "FK → inv_instrument_types.id",        "notes": ""},
            {"col": "serial_no",             "type": "VARCHAR(100)","nullable": "YES", "default": "",               "constraints": "",                                    "notes": ""},
            {"col": "manufacturer",          "type": "VARCHAR(200)","nullable": "YES", "default": "",               "constraints": "",                                    "notes": ""},
            {"col": "model",                 "type": "VARCHAR(200)","nullable": "YES", "default": "",               "constraints": "",                                    "notes": ""},
            {"col": "location",              "type": "VARCHAR(200)","nullable": "YES", "default": "",               "constraints": "",                                    "notes": ""},
            {"col": "purchase_date",         "type": "DATE",        "nullable": "YES", "default": "",               "constraints": "",                                    "notes": ""},
            {"col": "last_calibration_date", "type": "DATE",        "nullable": "YES", "default": "",               "constraints": "",                                    "notes": ""},
            {"col": "calibration_due_date",  "type": "DATE",        "nullable": "YES", "default": "",               "constraints": "",                                    "notes": ""},
            {"col": "calibration_status",    "type": "VARCHAR(20)", "nullable": "NO",  "default": "'OK'",           "constraints": "NOT NULL",                            "notes": "OK | DUE | OVERDUE | EXPIRED"},
            {"col": "status",                "type": "VARCHAR(30)", "nullable": "NO",  "default": "'ACTIVE'",       "constraints": "NOT NULL",                            "notes": "ACTIVE | INACTIVE | UNDER_CALIBRATION | DECOMMISSIONED"},
            {"col": "is_active",             "type": "BOOLEAN",     "nullable": "NO",  "default": "TRUE",           "constraints": "NOT NULL",                            "notes": ""},
        ],
    },

    # ── 16. inv_column_catalogue ──────────────────────────────────────────────
    {
        "name": "inv_column_catalogue",
        "description": "Chromatography column register with injection count tracking.",
        "group": "Catalogues",
        "columns": [
            {"col": "id",                    "type": "INTEGER",     "nullable": "NO",  "default": "auto-increment", "constraints": "PRIMARY KEY, INDEX",            "notes": ""},
            {"col": "column_id",             "type": "VARCHAR(100)","nullable": "NO",  "default": "",               "constraints": "UNIQUE, NOT NULL, INDEX",       "notes": "e.g. COL-001"},
            {"col": "name",                  "type": "VARCHAR(255)","nullable": "NO",  "default": "",               "constraints": "NOT NULL",                     "notes": ""},
            {"col": "column_type_id",        "type": "INTEGER",     "nullable": "YES", "default": "",               "constraints": "FK → inv_column_types.id",     "notes": ""},
            {"col": "serial_no",             "type": "VARCHAR(100)","nullable": "YES", "default": "",               "constraints": "",                             "notes": ""},
            {"col": "manufacturer",          "type": "VARCHAR(200)","nullable": "YES", "default": "",               "constraints": "",                             "notes": ""},
            {"col": "part_no",               "type": "VARCHAR(100)","nullable": "YES", "default": "",               "constraints": "",                             "notes": "Part / catalogue number"},
            {"col": "purchased_date",        "type": "DATE",        "nullable": "YES", "default": "",               "constraints": "",                             "notes": ""},
            {"col": "max_injections",        "type": "INTEGER",     "nullable": "YES", "default": "500",            "constraints": "",                             "notes": "Rated injection limit"},
            {"col": "cumulative_injections", "type": "INTEGER",     "nullable": "NO",  "default": "0",              "constraints": "NOT NULL",                     "notes": "Running injection count"},
            {"col": "status",                "type": "VARCHAR(20)", "nullable": "NO",  "default": "'ACTIVE'",       "constraints": "NOT NULL",                     "notes": "ACTIVE | INACTIVE | EXHAUSTED | RETIRED"},
            {"col": "is_active",             "type": "BOOLEAN",     "nullable": "NO",  "default": "TRUE",           "constraints": "NOT NULL",                     "notes": ""},
        ],
    },

    # ── 17. inv_maintenance_schedules ─────────────────────────────────────────
    {
        "name": "inv_maintenance_schedules",
        "description": "Planned maintenance schedule entries for equipment.",
        "group": "Schedules",
        "columns": [
            {"col": "id",               "type": "INTEGER",     "nullable": "NO",  "default": "auto-increment", "constraints": "PRIMARY KEY, INDEX",                      "notes": ""},
            {"col": "equipment_id",     "type": "INTEGER",     "nullable": "NO",  "default": "",               "constraints": "FK → inv_equipment_catalogue.id (CASCADE), NOT NULL","notes": ""},
            {"col": "maintenance_type", "type": "VARCHAR(100)","nullable": "YES", "default": "",               "constraints": "",                                        "notes": "Preventive | Corrective | Annual | Quarterly | etc."},
            {"col": "scheduled_date",   "type": "DATE",        "nullable": "NO",  "default": "",               "constraints": "NOT NULL",                                "notes": ""},
            {"col": "completed_date",   "type": "DATE",        "nullable": "YES", "default": "",               "constraints": "",                                        "notes": "Set on completion"},
            {"col": "technician",       "type": "VARCHAR(200)","nullable": "YES", "default": "",               "constraints": "",                                        "notes": ""},
            {"col": "status",           "type": "VARCHAR(20)", "nullable": "NO",  "default": "'DUE'",          "constraints": "NOT NULL",                                "notes": "DUE | IN_PROGRESS | COMPLETED | CANCELLED"},
            {"col": "notes",            "type": "TEXT",        "nullable": "YES", "default": "",               "constraints": "",                                        "notes": ""},
        ],
    },

    # ── 18. inv_calibration_schedules ─────────────────────────────────────────
    {
        "name": "inv_calibration_schedules",
        "description": "Calibration schedule entries for analytical instruments.",
        "group": "Schedules",
        "columns": [
            {"col": "id",               "type": "INTEGER",     "nullable": "NO",  "default": "auto-increment", "constraints": "PRIMARY KEY, INDEX",                        "notes": ""},
            {"col": "instrument_id",    "type": "INTEGER",     "nullable": "NO",  "default": "",               "constraints": "FK → inv_instrument_catalogue.id (CASCADE), NOT NULL","notes": ""},
            {"col": "calibration_type", "type": "VARCHAR(100)","nullable": "YES", "default": "",               "constraints": "",                                          "notes": "Internal | External | OQ | IQ | PQ | Annual | etc."},
            {"col": "scheduled_date",   "type": "DATE",        "nullable": "NO",  "default": "",               "constraints": "NOT NULL",                                  "notes": ""},
            {"col": "completed_date",   "type": "DATE",        "nullable": "YES", "default": "",               "constraints": "",                                          "notes": "Set on completion"},
            {"col": "technician",       "type": "VARCHAR(200)","nullable": "YES", "default": "",               "constraints": "",                                          "notes": ""},
            {"col": "certificate_no",   "type": "VARCHAR(100)","nullable": "YES", "default": "",               "constraints": "",                                          "notes": "Calibration certificate number"},
            {"col": "status",           "type": "VARCHAR(20)", "nullable": "NO",  "default": "'DUE'",          "constraints": "NOT NULL",                                  "notes": "DUE | IN_PROGRESS | COMPLETED | CANCELLED"},
            {"col": "notes",            "type": "TEXT",        "nullable": "YES", "default": "",               "constraints": "",                                          "notes": ""},
        ],
    },

    # ── 19. inv_equipment_verifications ───────────────────────────────────────
    {
        "name": "inv_equipment_verifications",
        "description": "Equipment qualification / periodic verification requests.",
        "group": "Verifications",
        "columns": [
            {"col": "id",           "type": "INTEGER",     "nullable": "NO",  "default": "auto-increment", "constraints": "PRIMARY KEY, INDEX",                          "notes": ""},
            {"col": "request_no",   "type": "VARCHAR(100)","nullable": "NO",  "default": "",               "constraints": "UNIQUE, NOT NULL, INDEX",                     "notes": "e.g. EQV-2025-001"},
            {"col": "equipment_id", "type": "INTEGER",     "nullable": "NO",  "default": "",               "constraints": "FK → inv_equipment_catalogue.id (CASCADE), NOT NULL","notes": ""},
            {"col": "requested_by", "type": "VARCHAR(200)","nullable": "YES", "default": "",               "constraints": "",                                            "notes": ""},
            {"col": "requested_at", "type": "TIMESTAMPTZ", "nullable": "YES", "default": "now()",          "constraints": "SERVER DEFAULT",                              "notes": ""},
            {"col": "verified_by",  "type": "VARCHAR(200)","nullable": "YES", "default": "",               "constraints": "",                                            "notes": ""},
            {"col": "verified_at",  "type": "TIMESTAMPTZ", "nullable": "YES", "default": "",               "constraints": "",                                            "notes": ""},
            {"col": "status",       "type": "VARCHAR(20)", "nullable": "NO",  "default": "'PENDING'",      "constraints": "NOT NULL",                                    "notes": "PENDING | VERIFIED | REJECTED"},
            {"col": "remarks",      "type": "TEXT",        "nullable": "YES", "default": "",               "constraints": "",                                            "notes": ""},
        ],
    },

    # ── 20. inv_instrument_verifications ──────────────────────────────────────
    {
        "name": "inv_instrument_verifications",
        "description": "Instrument qualification / periodic verification requests.",
        "group": "Verifications",
        "columns": [
            {"col": "id",            "type": "INTEGER",     "nullable": "NO",  "default": "auto-increment", "constraints": "PRIMARY KEY, INDEX",                            "notes": ""},
            {"col": "request_no",    "type": "VARCHAR(100)","nullable": "NO",  "default": "",               "constraints": "UNIQUE, NOT NULL, INDEX",                       "notes": "e.g. INV-2025-001"},
            {"col": "instrument_id", "type": "INTEGER",     "nullable": "NO",  "default": "",               "constraints": "FK → inv_instrument_catalogue.id (CASCADE), NOT NULL","notes": ""},
            {"col": "requested_by",  "type": "VARCHAR(200)","nullable": "YES", "default": "",               "constraints": "",                                              "notes": ""},
            {"col": "requested_at",  "type": "TIMESTAMPTZ", "nullable": "YES", "default": "now()",          "constraints": "SERVER DEFAULT",                                "notes": ""},
            {"col": "verified_by",   "type": "VARCHAR(200)","nullable": "YES", "default": "",               "constraints": "",                                              "notes": ""},
            {"col": "verified_at",   "type": "TIMESTAMPTZ", "nullable": "YES", "default": "",               "constraints": "",                                              "notes": ""},
            {"col": "status",        "type": "VARCHAR(20)", "nullable": "NO",  "default": "'PENDING'",      "constraints": "NOT NULL",                                      "notes": "PENDING | VERIFIED | REJECTED"},
            {"col": "remarks",       "type": "TEXT",        "nullable": "YES", "default": "",               "constraints": "",                                              "notes": ""},
        ],
    },

    # ── 21. inv_audit_trail ───────────────────────────────────────────────────
    {
        "name": "inv_audit_trail",
        "description": "System-wide immutable audit log for all inventory operations.",
        "group": "Audit",
        "columns": [
            {"col": "id",           "type": "INTEGER",     "nullable": "NO",  "default": "auto-increment", "constraints": "PRIMARY KEY, INDEX",    "notes": ""},
            {"col": "event_type",   "type": "VARCHAR(100)","nullable": "NO",  "default": "",               "constraints": "NOT NULL, INDEX",        "notes": "MATERIAL_CREATED | BATCH_RECEIVED | BATCH_ISSUED | STATUS_CHANGED | etc."},
            {"col": "entity_type",  "type": "VARCHAR(50)", "nullable": "NO",  "default": "",               "constraints": "NOT NULL",               "notes": "material | batch | stock_request | equipment | instrument | column"},
            {"col": "entity_id",    "type": "INTEGER",     "nullable": "YES", "default": "",               "constraints": "",                       "notes": "Primary key of the affected record"},
            {"col": "entity_ref",   "type": "VARCHAR(200)","nullable": "YES", "default": "",               "constraints": "",                       "notes": "Human-readable ref (batch_no, code, request_no)"},
            {"col": "performed_by", "type": "VARCHAR(200)","nullable": "YES", "default": "",               "constraints": "",                       "notes": "Username"},
            {"col": "performed_at", "type": "TIMESTAMPTZ", "nullable": "YES", "default": "now()",          "constraints": "SERVER DEFAULT, INDEX",  "notes": ""},
            {"col": "old_value",    "type": "TEXT",        "nullable": "YES", "default": "",               "constraints": "",                       "notes": "JSON snapshot before change"},
            {"col": "new_value",    "type": "TEXT",        "nullable": "YES", "default": "",               "constraints": "",                       "notes": "JSON snapshot after change"},
            {"col": "details",      "type": "TEXT",        "nullable": "YES", "default": "",               "constraints": "",                       "notes": "Free-text context"},
        ],
    },
]


# ─── Excel generation ─────────────────────────────────────────────────────────

GROUP_COLORS = {
    "Materials":        "D1FAE5",
    "Manufacturers":    "DBEAFE",
    "Batches":          "FEF9C3",
    "Stock Requests":   "FFE4E6",
    "Equipment Master": "EDE9FE",
    "Catalogues":       "FEF3C7",
    "Schedules":        "CCFBF1",
    "Verifications":    "FCE7F3",
    "Audit":            "F5F5F4",
}


def build_excel(output_path: str):
    wb = Workbook()

    # ── Cover / TOC sheet ─────────────────────────────────────────────────────
    ws_toc = wb.active
    ws_toc.title = "Table of Contents"

    # Title block
    ws_toc.merge_cells("A1:G1")
    c = ws_toc["A1"]
    c.value = "Inventory Master — Database Schema Reference"
    c.font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    c.fill = header_fill(TEAL_DARK)
    c.alignment = center_align()
    ws_toc.row_dimensions[1].height = 32

    ws_toc.merge_cells("A2:G2")
    c = ws_toc["A2"]
    c.value = "Database: chemia_eln  ·  Schema: public  ·  21 tables  ·  Generated from SQLAlchemy models"
    c.font = Font(name="Calibri", italic=True, size=10, color="44403C")
    c.fill = header_fill("F5F5F4")
    c.alignment = center_align()
    ws_toc.row_dimensions[2].height = 18

    ws_toc.row_dimensions[3].height = 8  # spacer

    # TOC column headers
    hdr_row = 4
    headers = ["#", "Table Name", "Group", "Description", "Columns", "FK Refs", "Sheet Link"]
    col_widths_toc = [5, 32, 18, 52, 10, 10, 16]

    for ci, (h, w) in enumerate(zip(headers, col_widths_toc), 1):
        cell = ws_toc.cell(row=hdr_row, column=ci, value=h)
        cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        cell.fill = header_fill(TEAL_MID)
        cell.alignment = center_align()
        cell.border = thin_border()
        ws_toc.column_dimensions[get_column_letter(ci)].width = w

    ws_toc.row_dimensions[hdr_row].height = 20

    # TOC rows — one per table
    for ti, tbl in enumerate(SCHEMA):
        r = hdr_row + 1 + ti
        row_fill = header_fill(GROUP_COLORS.get(tbl["group"], WHITE))

        # Count FK columns
        fk_count = sum(1 for col in tbl["columns"] if "FK →" in col["constraints"])

        values = [
            ti + 1,
            tbl["name"],
            tbl["group"],
            tbl["description"],
            len(tbl["columns"]),
            fk_count,
            f"→ {tbl['name'][:20]}",
        ]

        for ci, val in enumerate(values, 1):
            cell = ws_toc.cell(row=r, column=ci, value=val)
            cell.fill = row_fill
            cell.border = thin_border()
            cell.font = Font(name="Calibri", size=9)
            if ci in (1, 5, 6):
                cell.alignment = center_align()
            else:
                cell.alignment = left_align(wrap=(ci == 4))

        ws_toc.row_dimensions[r].height = 18

    # ── One sheet per table ───────────────────────────────────────────────────
    COL_HEADERS = [
        "#", "Column Name", "Data Type", "Nullable",
        "Default Value", "Constraints", "Notes / Allowed Values",
    ]
    COL_WIDTHS   = [5, 26, 18, 10, 18, 42, 50]

    for tbl in SCHEMA:
        # Sheet name max 31 chars; strip inv_ prefix for brevity
        sheet_name = tbl["name"].replace("inv_", "")[:31]
        ws = wb.create_sheet(title=sheet_name)

        # ── Table title banner ────────────────────────────────────────────────
        ws.merge_cells("A1:G1")
        c = ws["A1"]
        c.value = tbl["name"]
        c.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
        c.fill = header_fill(TEAL_DARK)
        c.alignment = center_align()
        ws.row_dimensions[1].height = 26

        # Group + description
        ws.merge_cells("A2:G2")
        c = ws["A2"]
        c.value = f"Group: {tbl['group']}   ·   {tbl['description']}"
        c.font = Font(name="Calibri", italic=True, size=9, color="44403C")
        c.fill = header_fill(TEAL_LIGHT)
        c.alignment = left_align()
        ws.row_dimensions[2].height = 16

        ws.row_dimensions[3].height = 6  # spacer

        # ── Column headers ────────────────────────────────────────────────────
        for ci, (h, w) in enumerate(zip(COL_HEADERS, COL_WIDTHS), 1):
            cell = ws.cell(row=4, column=ci, value=h)
            cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
            cell.fill = header_fill(TEAL_MID)
            cell.alignment = center_align()
            cell.border = thin_border()
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[4].height = 20

        # ── Data rows ─────────────────────────────────────────────────────────
        for ri, col in enumerate(tbl["columns"], 1):
            r = 4 + ri

            # Choose row fill
            is_pk  = "PRIMARY KEY" in col["constraints"]
            is_fk  = "FK →" in col["constraints"]
            is_nn  = "NOT NULL" in col["constraints"] and not is_pk and not is_fk
            if is_pk:
                row_fill = header_fill(GREEN_LIGHT)
            elif is_fk:
                row_fill = header_fill(AMBER_LIGHT)
            elif is_nn:
                row_fill = header_fill("F0FDF4")
            else:
                row_fill = header_fill(STONE_50 if ri % 2 == 0 else WHITE)

            values = [
                ri,
                col["col"],
                col["type"],
                col["nullable"],
                col["default"] or "—",
                col["constraints"] or "—",
                col["notes"] or "—",
            ]

            for ci, val in enumerate(values, 1):
                cell = ws.cell(row=r, column=ci, value=val)
                cell.fill = row_fill
                cell.border = thin_border()
                cell.font = Font(
                    name="Calibri", size=9,
                    bold=(ci == 2 and is_pk),
                    color=("0F766E" if is_pk and ci == 2 else "292524"),
                )
                if ci == 1:
                    cell.alignment = center_align()
                elif ci in (3, 4):
                    cell.alignment = center_align(wrap=False)
                else:
                    cell.alignment = left_align(wrap=(ci in (6, 7)))

            ws.row_dimensions[r].height = 18 if not any(
                len(str(col.get(k, ""))) > 35 for k in ("constraints", "notes")
            ) else 32

        # ── Legend ────────────────────────────────────────────────────────────
        legend_row = 4 + len(tbl["columns"]) + 2
        ws.merge_cells(f"A{legend_row}:G{legend_row}")
        c = ws.cell(row=legend_row, column=1,
                    value="  Legend:  GREEN = Primary Key   |   YELLOW = Foreign Key   |   LIGHT GREEN = NOT NULL   |   WHITE/GREY = Optional")
        c.font = Font(name="Calibri", italic=True, size=8, color="78716C")
        c.alignment = left_align()

        # Freeze header rows
        ws.freeze_panes = "B5"

    # ── Summary sheet: all tables in one view ─────────────────────────────────
    ws_all = wb.create_sheet(title="All Tables Summary")

    ws_all.merge_cells("A1:H1")
    c = ws_all["A1"]
    c.value = "All Tables — Column Summary"
    c.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    c.fill = header_fill(TEAL_DARK)
    c.alignment = center_align()
    ws_all.row_dimensions[1].height = 28

    ALL_HDRS = ["Table", "Column", "Data Type", "Nullable", "Default", "Constraints", "Notes", "Group"]
    ALL_WIDTHS = [30, 24, 18, 10, 16, 42, 42, 16]
    for ci, (h, w) in enumerate(zip(ALL_HDRS, ALL_WIDTHS), 1):
        cell = ws_all.cell(row=2, column=ci, value=h)
        cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        cell.fill = header_fill(TEAL_MID)
        cell.alignment = center_align()
        cell.border = thin_border()
        ws_all.column_dimensions[get_column_letter(ci)].width = w
    ws_all.row_dimensions[2].height = 20

    r = 3
    for tbl in SCHEMA:
        grp_color = GROUP_COLORS.get(tbl["group"], WHITE)
        for col in tbl["columns"]:
            is_pk = "PRIMARY KEY" in col["constraints"]
            is_fk = "FK →" in col["constraints"]
            if is_pk:
                rf = header_fill(GREEN_LIGHT)
            elif is_fk:
                rf = header_fill(AMBER_LIGHT)
            else:
                rf = header_fill(grp_color)

            vals = [
                tbl["name"], col["col"], col["type"], col["nullable"],
                col["default"] or "—", col["constraints"] or "—",
                col["notes"] or "—", tbl["group"],
            ]
            for ci, val in enumerate(vals, 1):
                cell = ws_all.cell(row=r, column=ci, value=val)
                cell.fill = rf
                cell.border = thin_border()
                cell.font = Font(name="Calibri", size=9)
                cell.alignment = left_align(wrap=(ci in (6, 7)))
            ws_all.row_dimensions[r].height = 15
            r += 1

    ws_all.freeze_panes = "B3"
    ws_all.auto_filter.ref = f"A2:H{r - 1}"

    # ── Save ──────────────────────────────────────────────────────────────────
    wb.save(output_path)
    print(f"Saved: {output_path}")
    print(f"Sheets : {len(wb.sheetnames)}")
    print(f"Tables : {len(SCHEMA)}")
    print(f"Columns: {sum(len(t['columns']) for t in SCHEMA)}")


if __name__ == "__main__":
    import os
    out = os.path.join(os.path.dirname(__file__), "Inventory_Master_DB_Schema.xlsx")
    build_excel(out)
